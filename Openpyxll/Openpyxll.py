#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim:fileencoding=utf-8

from openpyxl import load_workbook, workbook
from datetime import datetime
from openpyxl import Workbook
from openpyxl.descriptors.base import DateTime
import os

class Group:
    group = None
    specialty = None
    profile = None
    ochn = None
    year = None
    thread = None
    amount = None
    def __init__(self,gr, Specialty, Profile, Ochn, Year, Thread, Amount):
        self.group = gr
        self.specialty = Specialty
        self.profile = Profile
        self.ochn = Ochn
        self.year = Year
        self.thread = Thread
        self.amount = Amount

def contains (a, b):
    for i in a:
        if b.group==i.group:
            return True
    return False

def at(a,b):
    for i in range(0,len(a)):
        if a[i].group==b.group:
            return i

groups = []

wb = load_workbook("../Исходные_данные/Список-студентов_(фрагмент).xlsx",read_only = True)
ws = wb['Лист1']
for row in ws.rows:
    if row[30].value and row[30].value!="Группа":
        amount = 1
        groupn = row[30].value
        specialty = row[32].value.split(" ")[0]
        profile = row[31].value.split(" ")[0]
        ochn = row[36].value
        year = int(datetime.today().year - int(row[29].value))
        thread = specialty + " " + profile + " " + str(year)
        if(ochn=="очное"):
            ochn = "очн"
        else:
            ochn = "заочн"
        group = Group(groupn,specialty,profile,ochn,year,thread,amount)
        if not contains(groups, group):
            groups.append(group)
        else:
            groups[at(groups,group)].amount+=1
        print(groups[-1].group)

workbook = Workbook()
sheet = workbook.active
sheet["A1"] = "Группа"
sheet["B1"] = "Поток"
sheet["C1"] = "Специальность/\nНаправление"
sheet["D1"] = "Специализация/\nПрофиль"
sheet["E1"] = "Очн/\nЗаочн"
sheet["F1"] = "Год набора"
sheet["G1"] = "Кол-во\nстудентов"

nrow = 2
for group in groups:
    sheet.cell(row = nrow, column = 1).value = group.group
    sheet.cell(row = nrow, column = 2).value = group.thread
    sheet.cell(row = nrow, column = 3).value = group.specialty
    sheet.cell(row = nrow, column = 4).value = group.profile
    sheet.cell(row = nrow, column = 5).value = group.ochn
    sheet.cell(row = nrow, column = 6).value = group.year
    sheet.cell(row = nrow, column = 7).value = group.amount
    nrow+=1
workbook.save(filename = "../Результаты/work_openpyxl.xlsx")
wb.close()